#!/usr/bin/env python3
"""Build the spreadsheet deliverables for the recent home-sale photo run."""

from __future__ import annotations

import argparse
import csv
import json
from datetime import UTC, datetime
from pathlib import Path
from textwrap import fill

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WORKFLOW_ROOT = Path("unova") / "home-sale-inventory-desk"
DEFAULT_RUN_ROOT = WORKFLOW_ROOT / "runs" / "20260426-005844-recent-10-day-resale"


INVENTORY_COLUMNS = [
    "Item ID",
    "Photo IDs",
    "Area",
    "Item",
    "Brand / Model",
    "Category",
    "Condition",
    "Priority",
    "Recommended Platform",
    "Suggested Retail Price",
    "Quick Sale Price",
    "Average Online Price",
    "Online Price Range",
    "Acceptable Minimum",
    "Source Names",
    "Source Links",
    "Listing Title",
    "Listing Description",
    "Status",
    "Needed Evidence",
    "Confidence",
    "Notes",
]

SCREENING_COLUMNS = [
    "Screen ID",
    "Photo IDs",
    "Classification",
    "Action",
    "Reason",
    "Notes",
]

SOURCE_COLUMNS = [
    "Source ID",
    "Name",
    "URL",
    "Used For",
    "Observed Detail",
]


SOURCES = [
    {
        "Source ID": "SRC001",
        "Name": "RYOBI P2109VNM eBay used listing",
        "URL": "https://www.ebay.com/itm/157445327479",
        "Used For": "RYOBI 18V blower",
        "Observed Detail": "Used tool-only blower listed around $44 plus shipping.",
    },
    {
        "Source ID": "SRC002",
        "Name": "RYOBI P2606B eBay product/listing",
        "URL": "https://www.ebay.com/p/12032770082",
        "Used For": "RYOBI 22 inch hedge trimmer",
        "Observed Detail": "Open-box/listed examples around $80-$95.",
    },
    {
        "Source ID": "SRC003",
        "Name": "Stinger WD2025 Home Depot",
        "URL": "https://www.homedepot.com/p/100021998",
        "Used For": "Stinger compact wet/dry vac",
        "Observed Detail": "New 2.5 gal compact wet/dry vac listed at $34.97.",
    },
    {
        "Source ID": "SRC004",
        "Name": "BLACK+DECKER pole saw Home Depot",
        "URL": "https://www.homedepot.com/b/Outdoors-Outdoor-Power-Equipment-Chainsaws-Pole-Saws-Electric-Pole-Saws/BLACK-DECKER/10-in/N-5yc1vZcg4cZe7cZ1z1bari",
        "Used For": "BLACK+DECKER electric pole saw",
        "Observed Detail": "Comparable 10 inch models shown around $111-$119 new.",
    },
    {
        "Source ID": "SRC005",
        "Name": "Frigidaire FFFU14F2QW Home Depot",
        "URL": "https://www.homedepot.com/p/205555907",
        "Used For": "Frigidaire/Electrolux upright freezer",
        "Observed Detail": "13.8 cu ft frost-free upright freezer model reference.",
    },
    {
        "Source ID": "SRC006",
        "Name": "Toro 42 inch TimeCutter new reference",
        "URL": "https://aeoutdoorpower.com/products/toro-timecutter-42-zero-turn-mower-75740",
        "Used For": "Toro TimeCutter mower",
        "Observed Detail": "Comparable new 42 inch TimeCutter reference at $2,999.",
    },
    {
        "Source ID": "SRC007",
        "Name": "Used Toro TimeCutter 42 inch listing",
        "URL": "https://www.ronmowers.com/used-toro-timecutter-42-zero-turn-42-inch-riding-lawnmower-is-for-sale/",
        "Used For": "Toro TimeCutter mower",
        "Observed Detail": "Used 42 inch TimeCutter reference listed at $2,000.",
    },
    {
        "Source ID": "SRC008",
        "Name": "Milwaukee M18 miter saw Home Depot category",
        "URL": "https://www.homedepot.com/b/Tools-Power-Tools-Saws-Miter-Saws/Milwaukee-M18/Tool-Only/N-5yc1vZc2d7Z1z17rdrZ1z23lsz",
        "Used For": "Milwaukee M18 FUEL miter saw",
        "Observed Detail": "Comparable M18 FUEL cordless miter saws are premium tool-only items.",
    },
    {
        "Source ID": "SRC009",
        "Name": "National brass cash register eBay reference",
        "URL": "https://www.ebay.com/itm/205532708623",
        "Used For": "Antique brass cash register",
        "Observed Detail": "Partial/as-is parts can list low; complete registers vary widely by model.",
    },
    {
        "Source ID": "SRC010",
        "Name": "Howard Miller grandfather clocks",
        "URL": "https://howardmiller.com/collections/grandfather-clocks?page=2%2F1000",
        "Used For": "Grandfather clock baseline",
        "Observed Detail": "New floor clocks are expensive; local used value is strongly brand/condition dependent.",
    },
    {
        "Source ID": "SRC011",
        "Name": "Used grandfather clock resale discussion",
        "URL": "https://www.reddit.com/r/clocks/comments/1j7ezbs",
        "Used For": "Grandfather clock local pricing risk",
        "Observed Detail": "Local used grandfather clocks can move much lower than new retail due to pickup/repair cost.",
    },
    {
        "Source ID": "SRC012",
        "Name": "Ivers & Pond piano reference",
        "URL": "https://www.concertpitchpiano.com/Ivers-Pond-Piano-Prices.html",
        "Used For": "Ivers & Pond upright piano",
        "Observed Detail": "Brand history/reference; used piano value depends on tuning, action, and moving burden.",
    },
    {
        "Source ID": "SRC013",
        "Name": "Concrete gargoyle Etsy market",
        "URL": "https://www.etsy.com/market/concrete_gargoyle_statues",
        "Used For": "Concrete gargoyle garden statues",
        "Observed Detail": "New/listed gargoyle statues range from small inexpensive pieces to large $300+ pieces.",
    },
    {
        "Source ID": "SRC014",
        "Name": "Decorative suit of armor estate sale",
        "URL": "https://www.estatesales.net/marketplace/items/1742737",
        "Used For": "Decorative armor statue",
        "Observed Detail": "Life-size decorative armor estate item listed at $195 local pickup.",
    },
    {
        "Source ID": "SRC015",
        "Name": "Alice in Wonderland framed art market",
        "URL": "https://www.framedart.com/popular-character-art/framed-alice-in-wonderland-art-c36263",
        "Used For": "Framed fantasy/Alice art",
        "Observed Detail": "New framed character-art examples frequently list around $160-$220.",
    },
    {
        "Source ID": "SRC016",
        "Name": "Garage workbench used market",
        "URL": "https://www.used.forsale/garage-workbench",
        "Used For": "Garage cabinet/workbench",
        "Observed Detail": "Used workbench listings vary widely; page showed average around $242.",
    },
    {
        "Source ID": "SRC017",
        "Name": "Garage cabinet Craigslist sample",
        "URL": "https://portland.craigslist.org/search/sss?query=garage+storage+cabinets",
        "Used For": "Garage cabinet/workbench",
        "Observed Detail": "Local cabinet examples appeared around $40-$125, larger systems much higher.",
    },
    {
        "Source ID": "SRC018",
        "Name": "Kids ride-on ATV Walmart market",
        "URL": "https://www.walmart.com/c/kp/kids-ride-battery-powered-atv-quad",
        "Used For": "Kids ride-on ATV",
        "Observed Detail": "New ride-on ATV toys commonly run from about $70 to $225+ depending voltage/features.",
    },
    {
        "Source ID": "SRC019",
        "Name": "Graco TurboBooster recall page",
        "URL": "https://recalls.gracobaby.com/products/CS16TRBBST",
        "Used For": "Child booster seats safety hold",
        "Observed Detail": "Used only to flag recall/expiration verification before any resale decision.",
    },
    {
        "Source ID": "SRC020",
        "Name": "SKILSAW circular saw eBay category",
        "URL": "https://www.ebay.com/b/SKILSAW-Circular-Saw-Circular-Saws/71307/bn_7112978527",
        "Used For": "SKILSAW corded circular saw",
        "Observed Detail": "Used circular saw listings vary by model and condition.",
    },
    {
        "Source ID": "SRC021",
        "Name": "Vintage hall tree / coat rack Etsy market",
        "URL": "https://www.etsy.com/market/vintage_hall_tree_coat_rack",
        "Used For": "Freestanding coat tree",
        "Observed Detail": "Used/vintage coat trees vary widely; local pickup should be priced below shipped decor listings.",
    },
]


def source_links(*source_ids: str) -> str:
    lookup = {source["Source ID"]: source for source in SOURCES}
    return " | ".join(f"{source_id}: {lookup[source_id]['URL']}" for source_id in source_ids)


def source_names(*source_ids: str) -> str:
    lookup = {source["Source ID"]: source for source in SOURCES}
    return "; ".join(f"{source_id} {lookup[source_id]['Name']}" for source_id in source_ids)


def row(
    item_id: str,
    photos: str,
    area: str,
    item: str,
    brand: str,
    category: str,
    condition: str,
    priority: str,
    platform: str,
    suggested: str,
    quick: str,
    average: str,
    price_range: str,
    minimum: str,
    source_ids: tuple[str, ...],
    title: str,
    description: str,
    status: str,
    evidence: str,
    confidence: str,
    notes: str = "",
) -> dict[str, str]:
    return {
        "Item ID": item_id,
        "Photo IDs": photos,
        "Area": area,
        "Item": item,
        "Brand / Model": brand,
        "Category": category,
        "Condition": condition,
        "Priority": priority,
        "Recommended Platform": platform,
        "Suggested Retail Price": suggested,
        "Quick Sale Price": quick,
        "Average Online Price": average,
        "Online Price Range": price_range,
        "Acceptable Minimum": minimum,
        "Source Names": source_names(*source_ids) if source_ids else "",
        "Source Links": source_links(*source_ids) if source_ids else "",
        "Listing Title": title,
        "Listing Description": description,
        "Status": status,
        "Needed Evidence": evidence,
        "Confidence": confidence,
        "Notes": notes,
    }


INVENTORY = [
    row(
        "HS001",
        "P039-P043",
        "Garage / tools",
        "Zero-turn riding mower",
        "Toro TimeCutter, exact deck/model not yet visible",
        "Lawn equipment",
        "Used, dusty, appears complete",
        "A",
        "Facebook Marketplace; Craigslist",
        "$1,700",
        "$1,200",
        "$2,000",
        "$1,200-$2,500",
        "$1,000",
        ("SRC006", "SRC007"),
        "Toro TimeCutter Zero-Turn Riding Mower",
        "Toro TimeCutter zero-turn mower. Used and garage-stored. Best listed after confirming start, drive, blade engagement, deck size, engine, and hours if available.",
        "Research first",
        "Exact model/deck size, hours, battery condition, start/cut video",
        "Medium",
        "Biggest upside item in the current batch.",
    ),
    row(
        "HS002",
        "P026-P029",
        "Garage / tools",
        "Cordless compound miter saw",
        "Milwaukee M18 FUEL, exact model unknown",
        "Power tool",
        "Used, dusty, blade present, full model not visible",
        "A",
        "Facebook Marketplace; eBay",
        "$325",
        "$240",
        "$350",
        "$250-$500",
        "$200",
        ("SRC008",),
        "Milwaukee M18 FUEL Cordless Miter Saw",
        "Milwaukee M18 FUEL cordless miter saw. Used garage condition. Premium tool pickup item; price depends heavily on exact model, battery/charger inclusion, and working test.",
        "Needs exact model",
        "Model plate, blade size, battery/charger status, test cut video",
        "Medium",
    ),
    row(
        "HS003",
        "P034-P035",
        "Garage / appliances",
        "Upright freezer",
        "Frigidaire/Electrolux FFFU14F2QWJ, 13.8 cu ft",
        "Appliance",
        "Used, exterior looks clean from photo",
        "A",
        "Facebook Marketplace",
        "$225",
        "$150",
        "$300",
        "$150-$450",
        "$125",
        ("SRC005",),
        "Frigidaire 13.8 cu ft Upright Freezer",
        "Frigidaire/Electrolux frost-free upright freezer, model FFFU14F2QWJ. Great garage freezer or overflow freezer. Local pickup only.",
        "Ready after test",
        "Photo of inside, confirm cold temp, dimensions, pickup access",
        "High",
    ),
    row(
        "HS004",
        "P013; P040-P041",
        "Garage / storage",
        "Garage cabinet/workbench",
        "Brand unknown",
        "Garage storage",
        "Used, functional-looking, needs dimensions",
        "A",
        "Facebook Marketplace; Craigslist",
        "$125",
        "$75",
        "$150",
        "$70-$250",
        "$50",
        ("SRC016", "SRC017"),
        "Garage Storage Cabinet and Workbench",
        "Garage cabinet/workbench with upper cabinets, lower storage, and wood work surface. Useful shop or garage storage. Local pickup only.",
        "Ready after dimensions",
        "Dimensions, material, whether contents are included",
        "Medium",
    ),
    row(
        "HS005",
        "P030-P033",
        "Garage / tools",
        "Cordless leaf blower",
        "RYOBI ONE+ 18V P2109VNM",
        "Outdoor power tool",
        "Used, model label visible",
        "B",
        "Facebook Marketplace; eBay",
        "$35",
        "$25",
        "$45",
        "$30-$65 shipped comps",
        "$20",
        ("SRC001",),
        "RYOBI ONE+ 18V Cordless Leaf Blower",
        "RYOBI ONE+ 18V cordless blower, model P2109VNM. Tool appears used. Best sold with tested battery/charger or bundled with the RYOBI hedge trimmer.",
        "Ready after test",
        "Confirm tool-only vs battery included and show it running",
        "High",
    ),
    row(
        "HS006",
        "P008-P010; P012; P030",
        "Garage / tools",
        "Cordless hedge trimmer",
        "RYOBI ONE+ 18V 22 inch hedge trimmer, likely P2606/P2606B",
        "Outdoor power tool",
        "Used, blade/tool/box photos visible",
        "B",
        "Facebook Marketplace; eBay",
        "$65",
        "$45",
        "$80",
        "$55-$95",
        "$35",
        ("SRC002",),
        "RYOBI ONE+ 18V 22 in. Hedge Trimmer",
        "RYOBI ONE+ 18V 22-inch cordless hedge trimmer. Used condition with box shown. Strong quick local sale if tested.",
        "Ready after test",
        "Confirm exact model, battery included, blade condition",
        "High",
    ),
    row(
        "HS007",
        "P022-P024",
        "Garage / tools",
        "RYOBI battery and charger",
        "RYOBI P107 18V lithium battery plus charger",
        "Battery / charger",
        "Used, charge health unknown",
        "B",
        "Bundle with RYOBI tools",
        "$30",
        "$20",
        "$25",
        "$15-$40",
        "$15",
        (),
        "RYOBI ONE+ 18V Battery and Charger",
        "RYOBI ONE+ 18V lithium battery and charger. Best bundled with the blower and hedge trimmer if the battery holds charge.",
        "Bundle",
        "Charge test and photo of charger model",
        "Medium",
    ),
    row(
        "HS008",
        "P019-P020",
        "Garage / tools",
        "Compact wet/dry vacuum",
        "Stinger WD2025-style 2.5 gal vac",
        "Shop vacuum",
        "Used, dusty, accessories unclear",
        "B",
        "Facebook Marketplace; bundle",
        "$25",
        "$15",
        "$35",
        "$20-$45",
        "$10",
        ("SRC003",),
        "Stinger Compact Wet/Dry Utility Vacuum",
        "Compact Stinger wet/dry utility vac. Useful for garage, car, and small cleanup jobs. Used condition.",
        "Ready after test",
        "Confirm powers on and list hose/filter/accessories",
        "High",
    ),
    row(
        "HS009",
        "P016; P021",
        "Garage / tools",
        "Corded circular saw",
        "SKILSAW / Skil, exact model unknown",
        "Power tool",
        "Used, dusty, blade/cord condition unknown",
        "B",
        "Facebook Marketplace; tool lot",
        "$35",
        "$20",
        "$40",
        "$20-$60",
        "$15",
        ("SRC020",),
        "SKILSAW Corded Circular Saw",
        "Used SKILSAW corded circular saw. Good garage tool lot item if cord, guard, and trigger all test properly.",
        "Needs detail",
        "Model tag, blade size, guard/cord condition, working test",
        "Medium",
    ),
    row(
        "HS010",
        "P003; P014-P015; P025; P029",
        "Garage / yard",
        "Electric pole saw / chainsaw style yard tool",
        "BLACK+DECKER, exact model unknown",
        "Outdoor power tool",
        "Used, orange/black tool visible, model unclear",
        "B",
        "Facebook Marketplace; Craigslist",
        "$50",
        "$30",
        "$80",
        "$50-$120 new equivalents",
        "$20",
        ("SRC004",),
        "BLACK+DECKER Electric Pole Saw / Yard Saw",
        "BLACK+DECKER electric yard saw/pole saw style tool. Used condition. Good quick local pickup once the model and working condition are confirmed.",
        "Needs detail",
        "Full item photo, model tag, chain/blade condition, working test",
        "Medium",
    ),
    row(
        "HS011",
        "P017-P018",
        "Garage / yard",
        "Boxed Toro outdoor power tool",
        "Toro, exact model/contents unclear",
        "Outdoor power tool",
        "Box visible, contents not proven",
        "A",
        "Facebook Marketplace; Craigslist",
        "$90",
        "$50",
        "$90",
        "$50-$150",
        "$35",
        ("SRC006",),
        "Toro Outdoor Power Tool in Box",
        "Toro boxed outdoor power tool. Likely a yard-care item; do not post until the box model and actual contents are photographed.",
        "Research first",
        "Clear front/side box label, contents, working test",
        "Low",
    ),
    row(
        "HS012",
        "P001; P007; P011",
        "Garage / yard",
        "Yard tool lot",
        "Mixed rakes, shovels, handles, and extension cord",
        "Yard tools",
        "Used mixed lot",
        "C",
        "Facebook Marketplace",
        "$45",
        "$25",
        "$45",
        "$25-$75",
        "$15",
        (),
        "Yard Tool Lot with Extension Cord",
        "Mixed yard tool lot with rakes/shovels/handles and yellow extension cord. Best as one quick pickup bundle.",
        "Ready bundle",
        "Count tools and cord length/gauge if possible",
        "Medium",
    ),
    row(
        "HS013",
        "P037-P038",
        "Garage / kids",
        "Kids bicycle",
        "Red kids bike, brand unknown",
        "Bike",
        "Used, tires/brakes not verified",
        "B",
        "Facebook Marketplace",
        "$35",
        "$20",
        "$35",
        "$20-$60",
        "$10",
        (),
        "Red Kids Bike",
        "Used red kids bike. Quick local pickup item if tires hold air and brakes work.",
        "Ready after test",
        "Wheel size, tire/brake condition",
        "Medium",
    ),
    row(
        "HS014",
        "P044",
        "Garage / kids",
        "Kids ride-on ATV toy",
        "Brand/model unknown",
        "Ride-on toy",
        "Used, battery/charger unknown",
        "B",
        "Facebook Marketplace",
        "$50",
        "$30",
        "$100",
        "$70-$225 new equivalents",
        "$20",
        ("SRC018",),
        "Kids Ride-On ATV Toy",
        "Kids ride-on ATV style toy. Good quick local pickup item if battery and charger work.",
        "Needs test",
        "Brand/model, battery/charger, working video",
        "Medium",
    ),
    row(
        "HS015",
        "P004-P006; P036",
        "Garage / kids",
        "Child booster / car seats",
        "Graco visible",
        "Child safety gear",
        "Used, expiration/accident history unknown",
        "D",
        "Hold; do not list until verified",
        "$0",
        "$0",
        "$0-$20 only if safe/legal",
        "$0-$20",
        "$0",
        ("SRC019",),
        "Graco Booster / Car Seat",
        "Hold. Do not list until expiration, recall status, accident-free history, cleanliness, and marketplace policy are verified.",
        "Hold",
        "Expiration/manufacture date, recall check, accident-free confirmation",
        "Medium",
        "Safety-sensitive item; do not treat as a normal resale item.",
    ),
    row(
        "HS016",
        "P053",
        "Decor / collectibles",
        "Vintage-style glass jar",
        "San Miguel / recycled glass marking",
        "Decor",
        "Used",
        "C",
        "Facebook Marketplace; decor bundle",
        "$15",
        "$8",
        "$15",
        "$8-$25",
        "$5",
        (),
        "Vintage-Style Glass Jar",
        "Decorative glass jar with San Miguel recycled glass marking. Best bundled with other small decor.",
        "Bundle",
        "Full exterior photo and height",
        "Medium",
    ),
    row(
        "HS017",
        "P054",
        "Decor / furniture",
        "Freestanding coat tree",
        "Black metal tree-style coat rack",
        "Furniture",
        "Used, appears intact",
        "B",
        "Facebook Marketplace",
        "$45",
        "$25",
        "$50",
        "$25-$80",
        "$20",
        ("SRC021",),
        "Black Metal Coat Tree",
        "Black metal freestanding coat tree with branch-style hooks. Good entryway or bedroom piece.",
        "Ready after dimensions",
        "Height, stability, any missing hooks",
        "Medium",
    ),
    row(
        "HS018",
        "P055-P061; P115; P117; P121",
        "Decor / wall decor",
        "Gothic wall plaques and shields",
        "Mixed fantasy/gothic decor, brand unknown",
        "Wall decor",
        "Used, multiple pieces",
        "B",
        "Facebook Marketplace; eBay for special pieces",
        "$150 bundle",
        "$80 bundle",
        "$180 bundle",
        "$25-$80 per notable piece",
        "$60 bundle",
        ("SRC013",),
        "Gothic Wall Plaque and Shield Decor Lot",
        "Mixed gothic/fantasy wall decor lot including plaques, shields, and sculptural wall pieces. Strong themed-room bundle.",
        "Ready bundle",
        "Piece count, dimensions, closeups of marks/signatures",
        "Medium",
    ),
    row(
        "HS019",
        "P062-P063; P075-P080; P103-P113; P116; P120",
        "Decor / art",
        "Framed fantasy, Disney-style, and character art lot",
        "Mixed artists/brands; several signatures need closeups",
        "Wall art",
        "Used, framed pieces visible",
        "A",
        "Facebook Marketplace; eBay if signed/limited",
        "$300 bundle",
        "$150 bundle",
        "$170 per new framed reference for comparable character art",
        "$25-$220 per piece depending artist/signature",
        "$100 bundle",
        ("SRC015",),
        "Framed Fantasy and Character Art Collection",
        "Large collection of framed fantasy, character, and gothic wall art. Sell as themed bundle for speed, or split signed/limited pieces after closeup research.",
        "Research first",
        "Artist signatures, edition numbers, frame sizes, back labels",
        "Medium",
        "Potential hidden upside if any prints are limited or signed.",
    ),
    row(
        "HS020",
        "P064-P074; P106; P119; P122",
        "Decor / collectibles",
        "Fantasy statues and figurines lot",
        "Mixed gothic/fantasy/dragon/knight pieces",
        "Collectibles",
        "Used display condition",
        "B",
        "Facebook Marketplace; eBay for standout pieces",
        "$225 bundle",
        "$120 bundle",
        "$200 bundle",
        "$15-$125 per piece depending brand/size",
        "$90 bundle",
        ("SRC013",),
        "Fantasy Statue and Figurine Collection",
        "Mixed fantasy decor collection including skull, dragon, knight, helmet, and gothic building/statue pieces. Good themed display lot.",
        "Ready bundle",
        "Count pieces, dimensions, underside maker marks",
        "Medium",
    ),
    row(
        "HS021",
        "P065; P067-P068; P071-P072",
        "Decor / collectibles",
        "Nightmare Before Christmas style display pieces",
        "Jack Skellington / gothic character decor, exact brand unknown",
        "Collectibles",
        "Used display condition",
        "B",
        "Facebook Marketplace; eBay",
        "$125 bundle",
        "$75 bundle",
        "$120 bundle",
        "$50-$175",
        "$50 bundle",
        (),
        "Nightmare-Style Gothic Character Decor Lot",
        "Gothic character decor lot with teapot/cup pieces and Jack-style figurines. Best sold together unless maker marks show collector value.",
        "Needs marks",
        "Maker marks, Disney/licensed labels, condition closeups",
        "Medium",
    ),
    row(
        "HS022",
        "P093-P098",
        "Outdoor / decor",
        "Concrete gargoyle garden statues",
        "Concrete/resin, exact maker unknown",
        "Garden decor",
        "Used outdoor condition",
        "A",
        "Facebook Marketplace",
        "$180 pair/lot",
        "$100 pair/lot",
        "$150",
        "$70-$350 depending size/material",
        "$75",
        ("SRC013",),
        "Concrete Gargoyle Garden Statue Lot",
        "Concrete/resin gargoyle garden statue lot. Heavy local pickup decor for porch, patio, garden, or gothic room.",
        "Ready after dimensions",
        "Count, height/weight, cracks/chips, material",
        "Medium",
    ),
    row(
        "HS023",
        "P124-P126",
        "Furniture / decor",
        "Grandfather clock",
        "Brand/model unknown",
        "Clock",
        "Used, working status unknown",
        "A",
        "Facebook Marketplace; estate sale; clock collector group",
        "$350",
        "$150",
        "$400",
        "$150-$800 local used unless premium maker",
        "$100",
        ("SRC010", "SRC011"),
        "Grandfather Clock",
        "Tall wooden grandfather clock with decorative case. Local pickup only. Price depends on maker, movement, chimes, and working condition.",
        "Research first",
        "Maker label, movement label, dimensions, chime/clock working video",
        "Medium",
        "Bulky clocks can be slow even when nice.",
    ),
    row(
        "HS024",
        "P127",
        "Decor / collectibles",
        "Decorative suit of armor",
        "Full-size or large armor display, exact maker unknown",
        "Statement decor",
        "Used display condition",
        "A",
        "Facebook Marketplace; estate sale",
        "$225",
        "$150",
        "$195",
        "$150-$500 depending size/material",
        "$125",
        ("SRC014",),
        "Decorative Suit of Armor Display",
        "Decorative suit of armor display piece. Strong statement item for game room, bar, theater room, or themed office. Local pickup.",
        "Needs dimensions",
        "Height, material, missing/damaged parts, base stability",
        "Medium",
    ),
    row(
        "HS025",
        "P128-P129",
        "Music / furniture",
        "Upright piano",
        "Ivers & Pond, Boston",
        "Musical instrument",
        "Used, tuning/action unknown",
        "C",
        "Facebook Marketplace; local music groups",
        "$200",
        "$0-$100",
        "$250",
        "$0-$500 depending tune/move burden",
        "$0",
        ("SRC012",),
        "Ivers & Pond Upright Piano",
        "Ivers & Pond upright piano, Boston. Local pickup and professional moving required. Best price depends on tuning, action, soundboard, and whether buyer handles moving.",
        "Needs appraisal/test",
        "Serial number, tuning status, soundboard/action condition, mover access",
        "Low",
        "Pianos often move slowly; free-to-good-home may be fastest.",
    ),
    row(
        "HS026",
        "P130-P132",
        "Collectibles / antiques",
        "Antique brass cash register",
        "National Cash Register style, exact model unknown",
        "Antique / collectible",
        "Used display condition, function unknown",
        "A",
        "eBay; Facebook Marketplace; antique dealers",
        "$650",
        "$350",
        "$750",
        "$300-$1,200+ depending model/completeness",
        "$250",
        ("SRC009",),
        "Antique Brass Cash Register",
        "Ornate antique brass cash register with marble base/plate. Potentially valuable; research exact model and working condition before underpricing.",
        "Research first",
        "Model/serial plate, drawer function, keys, receipt mechanism, maker marks",
        "Medium",
        "Do not quick-sell until exact model and completeness are checked.",
    ),
    row(
        "HS027",
        "P197",
        "Furniture / decor",
        "Small wood side table / cabinet",
        "Brand unknown",
        "Furniture",
        "Used, partial view only",
        "C",
        "Facebook Marketplace",
        "$45",
        "$25",
        "$45",
        "$25-$80",
        "$15",
        (),
        "Small Wood Side Table / Cabinet",
        "Small wood side table/cabinet with drawers. Needs a clean full photo before posting.",
        "Needs detail",
        "Full straight-on photo, dimensions, drawer condition",
        "Low",
    ),
]


SCREENING = [
    {
        "Screen ID": "SKIP001",
        "Photo IDs": "P045-P052",
        "Classification": "Private / non-sale",
        "Action": "Do not list",
        "Reason": "People, ID cards, maps, business cards, meme, and location-sensitive images.",
        "Notes": "Keep out of marketplace uploads.",
    },
    {
        "Screen ID": "SKIP002",
        "Photo IDs": "P081-P092; P123; P133-P144; P145-P192; P198-P205; P209-P211; P214-P218",
        "Classification": "Private / personal",
        "Action": "Do not list",
        "Reason": "People and personal/body photos are not sale inventory.",
        "Notes": "Excluded from listing drafts and research queue.",
    },
    {
        "Screen ID": "SKIP003",
        "Photo IDs": "P183; P201; P212-P213",
        "Classification": "Documents / location / contacts",
        "Action": "Do not list",
        "Reason": "Parcel/location images, certificate/document photo, and business cards.",
        "Notes": "Potentially sensitive information.",
    },
    {
        "Screen ID": "SKIP004",
        "Photo IDs": "P136-P139",
        "Classification": "Store shelf reference",
        "Action": "Skip",
        "Reason": "Retail shelf photos are not home inventory.",
        "Notes": "Could be useful only if matching snack/drink stock, otherwise ignore.",
    },
    {
        "Screen ID": "SKIP005",
        "Photo IDs": "P206-P208",
        "Classification": "Regulated substance",
        "Action": "Do not list",
        "Reason": "Cannabis/plant photos should not enter a general resale workflow.",
        "Notes": "No marketplace listing recommendation.",
    },
    {
        "Screen ID": "SKIP006",
        "Photo IDs": "P190-P191",
        "Classification": "Ambiguous",
        "Action": "Needs user call",
        "Reason": "Possible tablet/fan/table visible inside personal mirror photo; not enough item evidence.",
        "Notes": "Ask for clean item-only photos if these are meant to sell.",
    },
    {
        "Screen ID": "SKIP007",
        "Photo IDs": "P186; P193-P196",
        "Classification": "Incident / social / meme",
        "Action": "Skip",
        "Reason": "Car fire, social photos, and meme content are not home sale inventory.",
        "Notes": "No listing action.",
    },
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write inventory CSV/XLSX/listing drafts.")
    parser.add_argument("--run-root", default=str(DEFAULT_RUN_ROOT))
    return parser.parse_args()


def write_csv(path: Path, columns: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def write_listing_drafts(path: Path, rows: list[dict[str, str]]) -> None:
    lines = [
        "# Listing Drafts",
        "",
        "Drafted from recent Photos.app item candidates. Review item evidence before posting.",
        "",
    ]
    for item in rows:
        if item["Status"] in {"Hold", "Needs detail"}:
            continue
        if item["Priority"] == "D":
            continue
        lines.extend(
            [
                f"## {item['Item ID']} - {item['Listing Title']}",
                "",
                f"- Platform: {item['Recommended Platform']}",
                f"- List price: {item['Suggested Retail Price']}",
                f"- Quick sale price: {item['Quick Sale Price']}",
                f"- Acceptable minimum: {item['Acceptable Minimum']}",
                f"- Photo IDs: {item['Photo IDs']}",
                "",
                fill(item["Listing Description"], width=96),
                "",
                f"Before posting: {item['Needed Evidence']}",
                "",
            ]
        )
    path.write_text("\n".join(lines), encoding="utf-8")


def research_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    queue = []
    for item in rows:
        if item["Status"] in {"Research first", "Needs exact model", "Needs appraisal/test"}:
            queue.append(
                {
                    "Item ID": item["Item ID"],
                    "Item": item["Item"],
                    "Reason": item["Status"],
                    "Needed Evidence": item["Needed Evidence"],
                    "Estimated Upside": item["Online Price Range"],
                    "Best Venue": item["Recommended Platform"],
                    "Confidence": item["Confidence"],
                }
            )
    return queue


def write_xlsx(path: Path, inventory: list[dict[str, str]], screening: list[dict[str, str]], sources: list[dict[str, str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_specs = [
        ("Inventory", INVENTORY_COLUMNS, inventory),
        ("Screening Notes", SCREENING_COLUMNS, screening),
        ("Sources", SOURCE_COLUMNS, sources),
        (
            "Research Queue",
            ["Item ID", "Item", "Reason", "Needed Evidence", "Estimated Upside", "Best Venue", "Confidence"],
            research_rows(inventory),
        ),
    ]
    header_fill = PatternFill("solid", fgColor="1D6D5F")
    header_font = Font(color="FFFFFF", bold=True)
    for title, columns, rows in sheet_specs:
        sheet = workbook.create_sheet(title)
        sheet.append(columns)
        for row_data in rows:
            sheet.append([row_data.get(column, "") for column in columns])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(columns, start=1):
            max_len = len(column)
            for row_index in range(2, sheet.max_row + 1):
                value = sheet.cell(row=row_index, column=index).value
                if value:
                    max_len = max(max_len, min(len(str(value)), 70))
                sheet.cell(row=row_index, column=index).alignment = Alignment(wrap_text=True, vertical="top")
            sheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), 72)
    workbook.save(path)


def write_status(path: Path, run_root: Path) -> None:
    counts = {
        "inventory_items": len(INVENTORY),
        "priority_a": sum(1 for item in INVENTORY if item["Priority"] == "A"),
        "priority_b": sum(1 for item in INVENTORY if item["Priority"] == "B"),
        "priority_c": sum(1 for item in INVENTORY if item["Priority"] == "C"),
        "priority_d": sum(1 for item in INVENTORY if item["Priority"] == "D"),
        "screening_rows": len(SCREENING),
        "research_queue_items": len(research_rows(INVENTORY)),
    }
    status = {
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "status": "spreadsheet_ready_first_pass",
        "run_root": str(run_root),
        "counts": counts,
        "artifact_paths": {
            "inventory_csv": str(run_root / "inventory-master.csv"),
            "workbook_xlsx": str(run_root / "inventory-master.xlsx"),
            "screening_csv": str(run_root / "screening-notes.csv"),
            "sources_csv": str(run_root / "sources.csv"),
            "research_queue_csv": str(run_root / "research-queue.csv"),
            "listing_drafts_md": str(run_root / "listing-drafts.md"),
        },
        "next_best_action": "Take exact model/inside/test photos for HS001, HS002, HS003, HS023, HS026.",
    }
    path.write_text(json.dumps(status, indent=2) + "\n", encoding="utf-8")


def write_summary(path: Path, run_root: Path) -> None:
    priority_a = [item for item in INVENTORY if item["Priority"] == "A"]
    fast = [item for item in INVENTORY if item["Status"] in {"Ready after test", "Ready bundle", "Ready after dimensions"}]
    research = research_rows(INVENTORY)
    lines = [
        "# Recent 10-Day Home Sale Inventory Summary",
        "",
        f"Generated: {datetime.now(UTC).replace(microsecond=0).isoformat()}",
        "",
        "## What Is Ready",
        "",
        f"- Spreadsheet workbook: `{run_root / 'inventory-master.xlsx'}`",
        f"- CSV inventory: `{run_root / 'inventory-master.csv'}`",
        f"- Listing drafts: `{run_root / 'listing-drafts.md'}`",
        f"- Review gallery: `{run_root / 'review-gallery.html'}`",
        "",
        "## Priority A Items",
        "",
    ]
    for item in priority_a:
        lines.append(
            f"- {item['Item ID']}: {item['Item']} - list around {item['Suggested Retail Price']} "
            f"(quick {item['Quick Sale Price']}); status: {item['Status']}."
        )
    lines.extend(["", "## Fastest Postable Items", ""])
    for item in fast[:10]:
        lines.append(
            f"- {item['Item ID']}: {item['Listing Title']} at {item['Suggested Retail Price']} "
            f"once evidence is checked."
        )
    lines.extend(["", "## Research Blockers", ""])
    for item in research:
        lines.append(f"- {item['Item ID']}: {item['Needed Evidence']}")
    lines.extend(
        [
            "",
            "## Safety / Privacy Gate",
            "",
            "- Private, people, documents, ID/contact, incident, and regulated-substance photos were screened out.",
            "- Child safety seats are held until expiration, recall, accident-free history, and platform policy are verified.",
            "",
            "## Tomorrow Morning Move",
            "",
            "Photograph exact model labels and working proof for HS001, HS002, HS003, HS023, and HS026, then post the ready B/C items while the high-value research finishes.",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    args = parse_args()
    run_root = Path(args.run_root).expanduser()
    if not run_root.is_absolute():
        run_root = Path.cwd() / run_root
    run_root.mkdir(parents=True, exist_ok=True)

    write_csv(run_root / "inventory-master.csv", INVENTORY_COLUMNS, INVENTORY)
    write_csv(run_root / "screening-notes.csv", SCREENING_COLUMNS, SCREENING)
    write_csv(run_root / "sources.csv", SOURCE_COLUMNS, SOURCES)
    write_csv(
        run_root / "research-queue.csv",
        ["Item ID", "Item", "Reason", "Needed Evidence", "Estimated Upside", "Best Venue", "Confidence"],
        research_rows(INVENTORY),
    )
    write_listing_drafts(run_root / "listing-drafts.md", INVENTORY)
    write_xlsx(run_root / "inventory-master.xlsx", INVENTORY, SCREENING, SOURCES)
    write_summary(run_root / "summary.md", run_root)
    write_status(run_root / "status.json", run_root)
    print(f"Inventory rows: {len(INVENTORY)}")
    print(f"Screening rows: {len(SCREENING)}")
    print(f"Workbook: {run_root / 'inventory-master.xlsx'}")
    print(f"CSV: {run_root / 'inventory-master.csv'}")


if __name__ == "__main__":
    main()
